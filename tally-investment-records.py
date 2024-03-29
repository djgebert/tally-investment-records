import PyPDF2
import openpyxl
from openpyxl.styles.numbers import BUILTIN_FORMATS
import textract
import re
from re import Match
import os
from typing import List
import sys
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import cell
from openpyxl.worksheet.dimensions import ColumnDimension
from collections import defaultdict
from datetime import datetime, timedelta
import os.path
from operator import attrgetter, truediv
import fiscalyear
from fiscalyear import FiscalDateTime
from progress.bar import ChargingBar

# We can handle the following record types as input (all pdfs)
WH_CONTRACTNOTE = 0
FAIR_DISTRIBUTION_ADVICE = 1
VDGR_REINVESTMENT_PLAN_ADVICE = 2

# When working with the spreadsheet, we will use these columns and formatting data
COLUMNS = {}
FORMAT = {}
idx = 1
for name, format_data in [
    ("Date", 17),
    ("Code", 9),
    ("Transaction type", 15),
    ("Quantity", 9),
    ("Subquantity", 11),
    ("Average price", 12),
    ("Brokerage", 10),
    ("Cost base", 11),
    ("Capital gain <= 1 year", 21),
    ("Capital gain > 1 year", 20),
    ("Net capital gain", 15),
    ("History", 50),
    ("Filename", 70)
]:
    COLUMNS[name] = idx
    idx += 1
    FORMAT[name] = format_data
del idx

# In Australia, the financial year begins on 1 July
# Fiscal year is represented by the year of its end date
fiscalyear.setup_fiscal_calendar('previous', 7, 1)

class InvestmentRecord():
    """Holds data for a single investment event, such as a sell or a buy.

    """

    def __init__(self, filename: str, record_type: int) -> None:
        self.filename = filename
        self.record_type = record_type
        self.populate()

    def populate_WH_ContractNote(self, text: str) -> None:
        # We can handle nabTrade Contract Notes
        # These will have a first line of text as follows
        if not text.startswith("WealthHub Securities Limited"):
            raise Exception("Unexpected format for WH_ContractNote file " + self.filename)

        self.trade_type = re.search(r"\n(.+) [Cc]onfirmation", text).group(1)
        # mFund transactions omit Trade date, so use a suitable substitute
        trade_date = re.search(r"Trade date:\n(.+)", text)
        if(trade_date):
            self.trade_date = trade_date.group(1)
        else:
            self.trade_date = re.search(r"As at date:\n(.+)", text).group(1)
        self.settlement_date = re.search(r"Settlement date:\n(.+)", text).group(1)
        self.confirmation_number = re.search(r"Confirmation number:\n(.+)", text).group(1)
        self.account_number = re.search(r"Account number:\n(.+)", text).group(1)
        self.hin = re.search(r"HIN:\n(.+)", text).group(1)
        # details holds multiple named groups in one big match
        details = re.search(
            r"Consideration\n(?P<quantity>.+)\n(?P<code>.+)\n(?P<security_description>[^$]+)(?P<average_price>.+)\n(?P<consideration>.+)\n", 
            text)
        self.quantity = details.group("quantity")
        self.code = details.group("code")
        self.average_price_per_share = details.group("average_price")
        self.brokerage = re.search(r"Brokerage\n(.+)", text).group(1)

        # Data type conversions for this record type
        self.trade_date = datetime.strptime(self.trade_date,"%d/%m/%Y")

    def populate_VDGR_Reinvestment_Plan_Advice(self, text: str) -> None:
        self.trade_type = "DRP"
        self.trade_date = re.search(r"Payment Date (.+)\r", text).group(1)
        self.quantity = re.search(r"Amount applied to (.+) ETF securities", text).group(1)
        '''self.code = re.search(r"ASX Code (.+)", text).group(1)'''
        self.code = "VDGR.ASX"
        self.average_price_per_share = re.search(r"securities allotted @ (.+) each", text).group(1)
        self.brokerage = "0"

        # Data type conversions for this record type
        self.trade_date = datetime.strptime(self.trade_date,"%d %B %Y")


    def populate_FAIR_Distribution_Advice(self, text: str) -> None:
        # These will have a first line of text as follows
        if not text.startswith("Class"):
            raise Exception("Unexpected format for FAIR_Distribution_Advice file " + self.filename)

        self.trade_type = "DRP"
        self.trade_date = re.search(r"Payment date:(.+)Record date:", text).group(1)
        self.quantity = re.search(r"This amount has been applied to (.+) units at ", text).group(1)
        self.code = re.search(r"ASX Code: (.+)Distribution Advice", text).group(1)
        self.average_price_per_share = re.search(r"units at (.+) per unit", text).group(1)
        self.brokerage = "0"

        # Data type conversions for this record type
        self.trade_date = datetime.strptime(self.trade_date,"%d %B %Y")

    def populate(self) -> None:
        """Populate instance variables from a pdf file.

        Args:
            filename: The pdf file containing data for population.
            
        """
        # Ensure requested file has a pdf extension
        if not self.filename.endswith(".pdf"):
            raise ValueError("Input file must have extension .pdf")
        
        if(self.record_type == VDGR_REINVESTMENT_PLAN_ADVICE):
            # Use OCR as PyPDF2 can't extract text from these
            # Dependencies required to be installed separately and put in PATH:
            #   tesseract-ocr-w64-setup-v5.0.0-alpha.20200328 
            #   poppler
            # Instructions:
            #   https://medium.com/quantrium-tech/installing-and-using-tesseract-4-on-windows-10-4f7930313f82 
            #   https://chadrick-kwag.net/install-poppler-in-windows/
            # Binaries:
            #   https://digi.bib.uni-mannheim.de/tesseract/
            #   https://github.com/oschwartz10612/poppler-windows/releases/
            # Copies of these binaries and instructions are also kept in the current repository's lib folder

            text = textract.process(self.filename, method='tesseract', language='eng').decode("utf-8")

        else:
            with open(self.filename, "rb") as f:
                pdf_reader = PyPDF2.PdfFileReader(f)

                # Ensure it's a single page
                if pdf_reader.numPages > 1:
                    raise Exception("Only single-page pdfs are accepted")

                # Get all the data we want
                text = pdf_reader.getPage(0).extractText()

        populators = {
            WH_CONTRACTNOTE: self.populate_WH_ContractNote,
            VDGR_REINVESTMENT_PLAN_ADVICE: self.populate_VDGR_Reinvestment_Plan_Advice,
            FAIR_DISTRIBUTION_ADVICE: self.populate_FAIR_Distribution_Advice
        }

        populators.get(self.record_type)(text)

        # Data type conversions for all record types
        self.quantity = float(self.quantity.replace(",", ""))
        self.average_price_per_share = float(self.average_price_per_share.replace("$",""))
        self.brokerage = float(self.brokerage.replace("$",""))


def get_investment_record_filenames(path: str) -> list:
    """Return a list of filenames matching the contract note format (WH_ContractNote_....pdf).

    Args:
        path: The path to search. All subdirectories within the path are also searched.

    """
    contract_note_filenames = []
    for dirpath, dirnames, filenames in os.walk(path):
        for filename in filenames:
            if re.fullmatch(r"WH_ContractNote_.+\.pdf", filename):
                record_type = WH_CONTRACTNOTE
            elif re.fullmatch(r"FAIR_Distribution_Advice_.+\.pdf", filename):
                record_type = FAIR_DISTRIBUTION_ADVICE
            elif re.fullmatch(r"VDGR_Reinvestment_Plan_Advice_.+\.pdf", filename):
                record_type = VDGR_REINVESTMENT_PLAN_ADVICE
            else:
                continue

            contract_note_filenames.append((os.path.join(dirpath, filename), record_type))
                
    return contract_note_filenames

def initialise_for_new_code(workbook: Workbook, code: str, records: list) -> Worksheet:
    # We can work on the workbook as well as the list of records in place
    sheet = workbook.create_sheet(code)
    for name, column_idx in COLUMNS.items(): 
        sheet.cell(1, column_idx).value = name

    records.sort(key=attrgetter("trade_date", "trade_type"))
    for record in records:
        if(record.trade_type.lower() in "buy", "mfund application", "drp"):
            record.available_quantity = record.quantity

    return sheet

def fiscal_year_check(sheet: Worksheet, fisc_year_start_idx: int, row_idx: int, prev_date: datetime, new_date: datetime, summaries: list, force_ytd_summary: bool) -> int:
    new_fiscal_year = FiscalDateTime(new_date.year, new_date.month, new_date.day).fiscal_year
    prev_fiscal_year = FiscalDateTime(prev_date.year, prev_date.month, prev_date.day).fiscal_year

    if (new_fiscal_year > prev_fiscal_year) or force_ytd_summary:

        sheet.cell(row_idx, COLUMNS["Date"]).value = "END OF FINANCIAL YEAR " + \
            str(prev_fiscal_year-1) + "-" + str(prev_fiscal_year)

        for column_name in ["Capital gain <= 1 year", "Capital gain > 1 year"]:
            col_letter = cell.get_column_letter(COLUMNS[column_name])
            formula = "=sum(" + col_letter + str(fisc_year_start_idx) + ":" \
                + col_letter + str(row_idx - 1) + ")"
            sheet.cell(row_idx, COLUMNS[column_name]).value = formula

        net_capital_gain_formula = "=" + cell.get_column_letter(COLUMNS["Capital gain <= 1 year"]) + str(row_idx) \
            + "+(" + cell.get_column_letter(COLUMNS["Capital gain > 1 year"]) + str(row_idx) +"/2)"
        sheet.cell(row_idx, COLUMNS["Net capital gain"]).value = net_capital_gain_formula

        for column_name in COLUMNS:
            sheet.cell(row_idx, COLUMNS[column_name]).style = "Accent1"

        # summaries.append({
        #     "fiscal_year_end": prev_fiscal_year, 
        #     "brokerage_ref": col_letter + str(row_idx),
        #     "capital_gain_ref": cap_gain_letter + str(row_idx),
        #     "is_ytd_only": force_ytd_summary
        #     })

        return row_idx + 1, row_idx + 1
    return row_idx, fisc_year_start_idx

def add_new_record_row(sheet: Worksheet, row_idx: int, record: InvestmentRecord) -> int:
    # Spreadsheet row is stored in the record object for later reference
    record.row_idx = row_idx
    sheet.cell(row_idx, COLUMNS["Date"]).value = record.trade_date
    sheet.cell(row_idx, COLUMNS["Code"]).value = record.code
    sheet.cell(row_idx, COLUMNS["Quantity"]).value = record.quantity
    sheet.cell(row_idx, COLUMNS["Average price"]).value = record.average_price_per_share
    sheet.cell(row_idx, COLUMNS["Transaction type"]).value = record.trade_type
    sheet.cell(row_idx, COLUMNS["Brokerage"]).value = record.brokerage
    sheet.cell(row_idx, COLUMNS["Filename"]).value = record.filename

def find_records_to_sell_fifo(records: list, sale_record: InvestmentRecord) -> list:
    # We assume records is already sorted
    quantity_to_sell = sale_record.quantity
    recs_and_quants_sold = []
    for record in records:
        if record is sale_record:
            raise Exception("Insufficient buy records found for sale record " + sale_record.filename)
        if record.trade_type.lower() in ("sell","mfund redemption"):
            continue
        if record.available_quantity >= quantity_to_sell:
            recs_and_quants_sold.append((record, quantity_to_sell))
            record.available_quantity -= quantity_to_sell
            quantity_to_sell = 0
            break
        elif record.available_quantity > 0:
            recs_and_quants_sold.append((record, record.available_quantity))
            quantity_to_sell -= record.available_quantity
            record.available_quantity = 0
    return recs_and_quants_sold

def add_sale_data(sheet: Worksheet, sale_record: InvestmentRecord, recs_and_quants_to_sell: list):
    """Returns the number of rows added for subquantities.
    
    """
    net_capital_gain_formula = "=(" + str(sale_record.quantity) + "*" \
    + cell.get_column_letter(COLUMNS["Average price"]) \
    + str(sale_record.row_idx) + ")"

    num_recs_and_quants = len(recs_and_quants_to_sell)
    using_subquantities = True if num_recs_and_quants > 1 else False
    current_row_idx = sale_record.row_idx + 1 if using_subquantities else sale_record.row_idx
    quantity_column_to_use = COLUMNS["Subquantity"] if using_subquantities else COLUMNS["Quantity"]

    for rec,quant in recs_and_quants_to_sell:

        if using_subquantities:
            sheet.cell(current_row_idx, quantity_column_to_use).value = quant
        # Else we assume the Quantity of the investment record was written earlier

        cost_base_formula = "=" + cell.get_column_letter(quantity_column_to_use) + str(current_row_idx) \
            + "*" + cell.get_column_letter(COLUMNS["Average price"]) + str(rec.row_idx) \
            + "+(" + cell.get_column_letter(COLUMNS["Brokerage"]) + str(rec.row_idx) \
            + "*" + cell.get_column_letter(quantity_column_to_use) + str(current_row_idx) \
            + "/" + cell.get_column_letter(COLUMNS["Quantity"]) + str(rec.row_idx) + ")" \
            + "+(" + cell.get_column_letter(COLUMNS["Brokerage"]) + str(sale_record.row_idx) \
            + "*" + cell.get_column_letter(quantity_column_to_use) + str(current_row_idx) \
            + "/" + cell.get_column_letter(COLUMNS["Quantity"]) + str(sale_record.row_idx) + ")"
        sheet.cell(current_row_idx, COLUMNS["Cost base"]).value = cost_base_formula

        capital_gain_formula = "=(" + cell.get_column_letter(quantity_column_to_use) + str(current_row_idx) \
            + "*" + cell.get_column_letter(COLUMNS["Average price"]) + str(sale_record.row_idx) \
            + ")-" + cell.get_column_letter(COLUMNS["Cost base"]) + str(current_row_idx) 
        if sale_record.trade_date > rec.trade_date + timedelta(days=365):
            capital_gain_column_to_use = COLUMNS["Capital gain > 1 year"]
        else:
            capital_gain_column_to_use = COLUMNS["Capital gain <= 1 year"]
        sheet.cell(current_row_idx, capital_gain_column_to_use).value = capital_gain_formula

        history = sheet.cell(rec.row_idx, COLUMNS["History"]).value
        new_history = "Sold " + str(quant) + " on " + sale_record.trade_date.strftime("%d/%m/%Y. ")
        sheet.cell(rec.row_idx, COLUMNS["History"]).value = (history + new_history if history else new_history)

        current_row_idx += 1

    return num_recs_and_quants if using_subquantities else 0

def format_code_sheet(sheet: Worksheet):
    for column_name in COLUMNS:
        sheet.cell(1, COLUMNS[column_name]).style = "Accent1"

    for column_name in ["Average price", "Brokerage", "Cost base", "Capital gain <= 1 year", "Capital gain > 1 year"]: 
        for this_cell in sheet[cell.get_column_letter(COLUMNS[column_name])]:
            pass # TODO: Set currency formatting here
    
    # openpyxl has trouble setting column width automatically, so we'll do it manually
    for col_name, col_width in FORMAT.items():
        sheet.column_dimensions[cell.get_column_letter(COLUMNS[col_name])].width = col_width
    

def add_summary_sheet(workbook: Workbook, all_fin_year_summaries: list):
    # Use the default sheet created with the workbook
    sheet = workbook["Sheet"]
    sheet.title = "Summary"
    
    # Expect a list of (string,[dict])
    #for code, summaries in all_fin_year_summaries:


def construct_investment_record_workbook(investment_records: List[InvestmentRecord]) -> Workbook:
    workbook = Workbook()

    # Normalise codes to just the ticker, removing any .ASX component.
    # This assumes everything we are trading is on the ASX. If we want to include other exchanges,
    # we will need to include the .ASX.
    for record in investment_records:
        record.code = re.split(r"\.", record.code)[0]

    investment_records_by_code = defaultdict(list)
    for record in investment_records:
        investment_records_by_code[record.code].append(record)
    all_fin_year_summaries = []
    row_idx = 1

    # We will have a sheet for each code
    for code in investment_records_by_code:
        records = investment_records_by_code[code]
        sheet = initialise_for_new_code(workbook, code, records)
        current_date = records[0].trade_date
        code_fin_year_summaries = []
        # Row 1 used for heading
        row_idx = 2
        fisc_year_start_idx = 2

        for record in records:
            row_idx, fisc_year_start_idx = fiscal_year_check(sheet, fisc_year_start_idx, row_idx, current_date, record.trade_date, code_fin_year_summaries, False)
            current_date = record.trade_date
            add_new_record_row(sheet, row_idx, record)
            if record.trade_type.lower() in ("buy", "drp"):
                cost_base_formula = "=" \
                    + cell.get_column_letter(COLUMNS["Quantity"]) \
                    + str(row_idx) + "*" \
                    + cell.get_column_letter(COLUMNS["Average price"]) \
                    + str(row_idx) + "+" \
                    + cell.get_column_letter(COLUMNS["Brokerage"]) \
                    + str(row_idx)                         
                sheet.cell(row_idx, COLUMNS["Cost base"]).value = cost_base_formula
            if record.trade_type.lower() in ("sell", "mfund redemption"):
                try:
                    recs_and_quants_sold = find_records_to_sell_fifo(records, record)
                except Exception as e:
                    sheet.cell(row_idx, COLUMNS["History"]).value = str(e)
                else:
                    pass
                    row_idx += add_sale_data(sheet, record, recs_and_quants_sold)
            row_idx += 1

        fiscal_year_check(sheet, fisc_year_start_idx, row_idx, current_date, current_date, code_fin_year_summaries, True)
        all_fin_year_summaries.append((code, code_fin_year_summaries))

        format_code_sheet(sheet)
    
    add_summary_sheet(workbook, all_fin_year_summaries)
    return workbook

def display_help():
    """Print a help message for this script to the terminal.
    """
    raise Exception("To be completed.")

def save_workbook(workbook: Workbook, filename: str):
    if os.path.isfile(filename):
        os.rename(filename, filename+datetime.now().strftime(".%Y-%m-%d.%H.%M.%S.bak"))    
    workbook.save(filename)    
 
if __name__ == "__main__":
    if(len(sys.argv) > 1):
            if(sys.argv[1] in ["-h", "help"]):
                display_help()
                exit()
            else:
                path_to_search = sys.argv[1]
    else:
        path_to_search = "."

    filenames = get_investment_record_filenames(path_to_search)
    progress_bar = ChargingBar('Processing', max=len(filenames))
    investment_records = []
    for filename, record_type in filenames:
        investment_records.append(InvestmentRecord(filename, record_type))
        progress_bar.next()
    progress_bar.finish()
    workbook = construct_investment_record_workbook(investment_records)
    save_workbook(workbook, "Investment_Record_Tally.xlsx")